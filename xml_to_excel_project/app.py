from flask import Flask, request, render_template, send_file
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from io import BytesIO

app = Flask(__name__)

# Helper function to strip namespace from an XML tag
def strip_namespace(tag):
    return tag.split('}')[-1] if '}' in tag else tag

# Append a row to the Excel sheet with appropriate indentation
def append_row(field_name, value, ws, depth):
    indented_field_name = '  ' * depth + field_name  # Add spaces for indentation
    ws.append([indented_field_name, 'String', '*', 'N', value])

# Generate the Excel file based on the XML payload
def generate_excel(xml_payload):
    root = ET.fromstring(xml_payload)
    wb = Workbook()
    ws = wb.active
    ws.title = "Response Data"
    
    # Define styles for header and cells
    header_font = Font(bold=True, color="FFFFFF")  # White text for headers
    header_fill = PatternFill(start_color="4B4B4B", end_color="4B4B4B", fill_type="solid")  # Dark grey background for headers
    cell_font = Font(color="FFFFFF")  # White text for content cells
    cell_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")  # Dark background for content cells
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))  # Thin borders
    
    # Header row
    headers = ["Field Name", "Type", "Length", "Mandatory", "Description"]
    ws.append(headers)
    
    # Apply header styles
    for col_num, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # Recursive function to process XML elements with indentation
    def process_element(element, depth):
        field_name = strip_namespace(element.tag)
        value = element.text.strip() if element.text else ""
        row = ['  ' * depth + field_name, "String", "*", "N", value]
        ws.append(row)
        
        # Apply styles to the row
        for cell in ws[ws.max_row]:
            cell.font = cell_font
            cell.fill = cell_fill
            cell.border = thin_border

        # Recursively process child elements
        for child in element:
            process_element(child, depth + 1)

    # Start processing from the root element
    process_element(root, 0)

    # Adjust column widths for better readability
    ws.column_dimensions['A'].width = 40  # Field Name
    ws.column_dimensions['B'].width = 10  # Type
    ws.column_dimensions['C'].width = 10  # Length
    ws.column_dimensions['D'].width = 10  # Mandatory
    ws.column_dimensions['E'].width = 30  # Description

    # Save the workbook to a BytesIO object
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# Extract a dynamic filename based on XML payload structure
def get_filename_from_xml(xml_payload):
    root = ET.fromstring(xml_payload)
    body = root.find('body')

    if body is not None:
        child_tags = [strip_namespace(child.tag) for child in body]
        if child_tags:
            return f"{child_tags[0]}.xlsx"
    
    return "response_data.xlsx"

# Define routes
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/generate', methods=['POST'])
def generate():
    xml_payload = request.form['xml_payload'].strip()  # Strip whitespace from input

    if not xml_payload:
        return "Error: XML payload is empty", 400

    try:
        # Check if the XML is well-formed
        ET.fromstring(xml_payload)

        # Generate the Excel file
        output = generate_excel(xml_payload)
        filename = get_filename_from_xml(xml_payload)
        return send_file(output, download_name=filename, as_attachment=True)
    except ET.ParseError as parse_error:
        return f"XML Parse Error: {str(parse_error)}", 400
    except Exception as e:
        return f"Error generating Excel: {str(e)}", 400

if __name__ == '__main__':
    app.run(debug=True)
